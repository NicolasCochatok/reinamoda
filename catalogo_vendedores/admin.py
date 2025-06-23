from django.contrib import admin, messages
from django.http import HttpResponse
from .models import Venta
import openpyxl
from openpyxl.utils import get_column_letter


@admin.register(Venta)
class VentaAdmin(admin.ModelAdmin):
    list_display = (
        'producto', 'cantidad', 'stock_actual', 'cliente',
        'metodo_pago', 'pagado', 'fecha_venta', 'vendedor'
    )
    list_editable = ('metodo_pago', 'pagado')
    list_filter = ('metodo_pago', 'pagado', 'fecha_venta')
    search_fields = (
        'cliente',
        'producto__product_name',
        'producto__codigo_unico',
        'producto__codigo_proveedor',
        'vendedor__email',
        'vendedor__username'
    )
    date_hierarchy = 'fecha_venta'
    actions = ['exportar_a_excel', 'marcar_como_pagadas']

    def stock_actual(self, obj):
        return obj.producto.stock
    stock_actual.short_description = 'Stock actual'

    def save_model(self, request, obj, form, change):
        if not obj.vendedor:
            obj.vendedor = request.user
        try:
            super().save_model(request, obj, form, change)
        except ValueError as e:
            self.message_user(request, f"Error: {e}", level=messages.ERROR)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter(vendedor=request.user)

    def has_change_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if obj is None or obj.vendedor == request.user:
            return True
        return False

    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if obj is None or obj.vendedor == request.user:
            return True
        return False

    def exportar_a_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"

        columnas = [
            "Producto", "Cantidad", "Cliente", "Método de pago",
            "Pagado", "Precio unitario", "Fecha", "Vendedor"
        ]
        for col_num, columna in enumerate(columnas, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = columna

        for fila_num, venta in enumerate(queryset, 2):
            datos = [
                venta.producto.product_name,
                venta.cantidad,
                venta.cliente,
                venta.metodo_pago,
                "Sí" if venta.pagado else "No",
                float(venta.precio_unitario),
                venta.fecha_venta.strftime('%Y-%m-%d %H:%M'),
                venta.vendedor.email if venta.vendedor else ''
            ]
            for col_num, valor in enumerate(datos, 1):
                ws.cell(row=fila_num, column=col_num, value=valor)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
        wb.save(response)
        return response

    exportar_a_excel.short_description = "📥 Exportar ventas seleccionadas a Excel"

    def marcar_como_pagadas(self, request, queryset):
        actualizadas = queryset.update(pagado=True)
        self.message_user(request, f"{actualizadas} venta(s) marcadas como pagadas.", level=messages.SUCCESS)

    marcar_como_pagadas.short_description = "✅ Marcar como pagadas"
